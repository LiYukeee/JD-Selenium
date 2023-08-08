import time

from openpyxl import load_workbook

import win32com.client as win32
import os
from url2CompanyName import url2Company, url2ShopName
from url2Message import url2Data
import re
from string import ascii_uppercase as alphabet
import configparser
from driverOperation import driverInit

config = configparser.ConfigParser()
configPath = r"./data/config.ini"
config.read(configPath)
waitTime = int(config['asd']['waitTime'])
restart = int(config['asd']['restart'])
restartWaitTime = int(config['asd']['restartWaitTime'])


def exchange(fold):
    """
    :param fold: 文件夹
    :return:
    """
    path = os.getcwd()
    path = os.path.join(path, fold)
    files = os.listdir(path)
    for file_name in files:
        if file_name.rsplit('.', 1)[-1] == 'xls':
            fname = os.path.join(path, file_name)
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            wb = excel.Workbooks.Open(fname)
            # 在原来的位置创建出：原名+'.xlsx'文件
            wb.SaveAs(fname + "x", FileFormat=51)  # FileFormat = 51 is for .xlsx extension
            wb.Close()  # FileFormat = 56 is for .xls extension
            excel.Application.Quit()
            os.remove(fname)
            return file_name.rsplit('.', 1)[0] + '.xlsx'
        if file_name.rsplit('.', 1)[-1] == 'xlsx':
            return file_name


def writeXlsx(workSheet, row, col, value):
    """
    写入表格，输入行和列，起始位置为第一行第一列\n
    :param workSheet: 列表对象
    :param row: 行
    :param col: 列
    :param value: 值
    :return:
    """
    colKey = alphabet[col - 1]
    rowKey = row
    key = "{}{}".format(colKey, rowKey)
    workSheet[key] = value


def readXlsx(workSheet, row, col):
    """
    读取表格，输入行和列，起始位置为第一行第一列
    :param workSheet: 列表对象
    :param row: 行
    :param col: 列
    :return: value
    """
    colKey = alphabet[col - 1]
    rowKey = row
    key = "{}{}".format(colKey, rowKey)
    return workSheet[key].value


def readData(workSheet, sheetDim):
    res = {}
    for i in range(2, sheetDim[0] + 1):
        url = readXlsx(workSheet, row=i, col=1)
        if url == None:
            continue
        if readXlsx(workSheet, row=i, col=sheetDim[1] + 4) == None:
            continue
        id = re.findall("""[1-9]\d*""", url)[0]
        pdfExists = readXlsx(workSheet, row=i, col=sheetDim[1] + 1)
        shopNameExists = readXlsx(workSheet, row=i, col=sheetDim[1] + 2)
        companyName = readXlsx(workSheet, row=i, col=sheetDim[1] + 3)
        shopName = readXlsx(workSheet, row=i, col=sheetDim[1] + 4)
        pdfPath = readXlsx(workSheet, row=i, col=sheetDim[1] + 5)
        res[id] = [pdfExists, shopNameExists, companyName, shopName, pdfPath]
    return res


def main():
    # 将目录下的xls转换为xlsx格式
    driver = driverInit(r"D:/")
    driver.set_window_size(1600, 1000)
    xlsxName = exchange('data')

    wb = load_workbook(r'./data/{}'.format(xlsxName))
    sheet = wb[wb.sheetnames[0]]  # 选定第一个表

    # 获取表格的尺寸
    sheetDim1 = int(re.findall("""[1-9]\d*""", sheet.dimensions)[-1])
    sheetDim2 = sheet.dimensions.split(":")[-1].replace(str(sheetDim1), "")
    sheetDim2 = alphabet.index(sheetDim2) + 1
    sheetDim = [sheetDim1, sheetDim2]

    # 写入表头
    if readXlsx(sheet, row=1, col=sheetDim[1]) == 'PDF路径' and readXlsx(sheet, row=1, col=sheetDim[1] - 1) == '店铺名':
        sheetDim[1] = sheetDim[1] - 5
    else:
        writeXlsx(sheet, row=1, col=sheetDim[1] + 1, value="是否获取PDF")
        writeXlsx(sheet, row=1, col=sheetDim[1] + 2, value="是否已经尝试提取店铺名")
        writeXlsx(sheet, row=1, col=sheetDim[1] + 3, value="公司名")
        writeXlsx(sheet, row=1, col=sheetDim[1] + 4, value="店铺名")
        writeXlsx(sheet, row=1, col=sheetDim[1] + 5, value="PDF路径")

    # 数据
    count = 0
    result = readData(sheet, sheetDim)
    for i in range(2, sheetDim[0] + 1):
        # 等待时间倍率
        waitTimeBase = 0
        if count == restart:
            count = 0
            print("重新启动..")
            driver.close()
            time.sleep(restartWaitTime)
            driver = driverInit(r"D:/")
            driver.set_window_size(1600, 1000)
        url = readXlsx(sheet, row=i, col=1)
        # 如果链接为空则进行下一条
        if url == None:
            continue
        id = re.findall("""[1-9]\d*""", url)[0]
        ##############################################
        # 判断是否重复爬虫
        data = ""
        if id in result.keys():  # 如果已经有重复数据
            data = result[id]
            writeXlsx(sheet, row=i, col=sheetDim[1] + 1, value=data[0])
            writeXlsx(sheet, row=i, col=sheetDim[1] + 2, value=data[1])
            writeXlsx(sheet, row=i, col=sheetDim[1] + 3, value=data[2])
            writeXlsx(sheet, row=i, col=sheetDim[1] + 4, value=data[3])
            writeXlsx(sheet, row=i, col=sheetDim[1] + 5, value=data[4])
            try:
                wb.save(r'./data/{}'.format(xlsxName))
            except:
                a = 1
            continue
        ##############################################
        # 没有存储此条数据，开始爬虫
        # 1 店铺名相关
        result[id] = [None, None, None, None, None]
        shopName = ""
        if readXlsx(sheet, row=i, col=sheetDim[1] + 4) is None:
            try:
                waitTimeBase = 1
                shopName = url2ShopName(driver, url)
                writeXlsx(sheet, row=i, col=sheetDim[1] + 4, value=shopName)
            except:
                shopName = ""
        else:
            shopName = readXlsx(sheet, row=i, col=sheetDim[1] + 4)
        result[id][4 - 1] = shopName
        # 2 pdf相关
        if readXlsx(sheet, row=i, col=sheetDim[1] + 1) != "是":
            count += 1
            waitTimeBase = 1
            try:
                pdfPath = url2Data(driver, url)
                writeXlsx(sheet, row=i, col=sheetDim[1] + 1, value="是")
                result[id][1 - 1] = "是"
                writeXlsx(sheet, row=i, col=sheetDim[1] + 5, value=pdfPath)
                result[id][5 - 1] = pdfPath
            except:
                writeXlsx(sheet, row=i, col=sheetDim[1] + 1, value="失败")
                result[id][1 - 1] = "失败"
        # 3 公司名相关
        if readXlsx(sheet, row=i, col=sheetDim[1] + 2) != "是" and "自营" not in shopName:
            waitTimeBase = 1
            try:
                companyName = url2Company(driver, url)
                writeXlsx(sheet, row=i, col=sheetDim[1] + 2, value="是")
                result[id][2 - 1] = "是"
                writeXlsx(sheet, row=i, col=sheetDim[1] + 3, value=companyName)
                result[id][3 - 1] = companyName
            except:
                writeXlsx(sheet, row=i, col=sheetDim[1] + 2, value="失败")
                result[id][2 - 1] = "失败"
        # 尝试保存，可中途打开，但不要修改其中内容
        try:
            wb.save(r'./data/{}'.format(xlsxName))
        except:
            a = 1
        time.sleep(waitTime * waitTimeBase)


if __name__ == '__main__':
    main()
    print("完成....")
    time.sleep(9999)
